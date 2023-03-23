import RPi.GPIO as GPIO
import serial
import time
from time import sleep
from datetime import datetime
import os
import numpy as np
import pandas as pd
import math
import threading
from openpyxl import Workbook
from openpyxl import load_workbook

# MQTT
from Adafruit_IO import Client, Feed

# for scheduling
from apscheduler.schedulers.background import BackgroundScheduler

# for machine learning purposed
import joblib
from sklearn.ensemble import RandomForestClassifier
from sklearn.neighbors import KNeighborsClassifier

# for GUI purposed
import tkinter as tk
from tkinter import *
from tkinter import ttk
import tkinter.font as tkFont
from tkinter import messagebox
from PIL import Image, ImageTk

# Change scheduler interval here !!!!
request_interval = 75
subscribe_interval = 6

# Change model usage, 0 = Random Forest, 1 = KNN
model = 1

# -------------------- MQTT adafruit IO --------------------
mqtt_Username = "ichsan27"
mqtt_Password = "aio_MEpE30o8v1AKaYMOW0zpjUAFsG2x"

# publish
humidity = "plant-slash-humidity"
lux = "plant-slash-lux"
temp = "plant-slash-temp"
update_condition = "plant-slash-result"
note = "plant-slash-note"

# subscribe
lamp_state = "plant-slash-onlamp"
pump_state = "plant-slash-onpump"
pump_onDuration = "plant-slash-pumponduration"
pump_offDuration = "plant-slash-pumpoffduration"

# initial value
condition_value = "Tidak Ideal"
note_value = "Kelembaban, Suhu air, dan insitas cahaya tidak ideal"

aio = Client(mqtt_Username, mqtt_Password)

# -------------------- GPIO Define -----------------------------
GPIO.setmode(GPIO.BCM) # Broadcom chip-spesific pin number

ser = serial.Serial('/dev/ttyS0', # replace tty50 with AM0 for Pi1, Pi2, Pi0
                     baudrate = 38400,
                     parity = serial.PARITY_NONE,
                     stopbits = serial.STOPBITS_ONE,
                     bytesize = serial.EIGHTBITS,
                     timeout = 1)

convert_char_to_int = {
    "\\t" : 9,
    "\\n" : 10,
    '\\r' : 13,
    ' ' : 32,
    '!' : 33,
    '"' : 34,
    '#' : 35,
    '$' : 36,
    '%' : 37,
    '&' : 38,
    "\'": 39,
    '(' : 40,
    ')' : 41,
    '*' : 42,
    '+' : 43,
    ',' : 44,
    '-' : 45,
    '.' : 46,
    '/' : 47,
    '0' : 48,
    '1' : 49,
    '2' : 50,
    '3' : 51,
    '4' : 52,
    '5' : 53,
    '6' : 54,
    '7' : 55,
    '8' : 56,
    '9' : 57,
    ':' : 58,
    ';' : 59,
    '<' : 60,
    '=' : 61,
    '>' : 62,
    '?' : 63,
    '@' : 64,
    'A' : 65,
    'B' : 66,
    'C' : 67,
    'D' : 68,
    'E' : 69,
    'F' : 70,
    'G' : 71,
    'H' : 72,
    'I' : 73,
    'J' : 74,
    'K' : 75,
    'L' : 76,
    'M' : 77,
    'N' : 78,
    'O' : 79,
    "P" : 80,
    'Q' : 81,
    'R' : 82,
    'S' : 83,
    'T' : 84,
    'U' : 85,
    'V' : 86,
    'W' : 87,
    'X' : 88,
    'Y' : 89,
    'Z' : 90,
    '[' : 91,
    '\\\\': 92,
    ']' : 93,
    '^' : 94,
    '_' : 95,
    "`" : 96,
    'a' : 97,
    'b' : 98,
    'c' : 99,
    'd' : 100,
    'e' : 101,
    'f' : 102,
    'g' : 103,
    'h' : 104,
    'i' : 105,
    'j' : 106,
    'k' : 107,
    'l' : 108,
    'm' : 109,
    'n' : 110,
    'o' : 111,
    'p' : 112,
    'q' : 113,
    'r' : 114,
    's' : 115,
    't' : 116,
    'u' : 117,
    'v' : 118,
    'w' : 119,
    'x' : 120,
    'y' : 121,
    'z' : 122,
    '{' : 123,
    '|' : 124,
    '}' : 125,
    '~' : 126,
}

# ---------------------------- Conversion data ---------------------
convert_res = {
    0 : ["Ideal",""],
    1 : ["Tidak Ideal", "Suhu air tidak ideal"],
    2 : ["Tidak Ideal", "Kelembapan udara tidak ideal"],
    3 : ["Tidak Ideal", "Intensitas cahaya tidak ideal"],
    4 : ["Tidak Ideal", "Suhu air dan kelembapan tidak ideal"],
    5 : ["Tidak Ideal", "Kelembapan dan intensitas cahaya tidak ideal"],
    6 : ["Tidak Ideal", "Suhu air dan intensitas cahaya tidak ideal"],
    7 : ["Tidak Ideal", "Suhu air, kelembapan, dan intensitas cahaya tidak ideal"],
}

res_to_out = {
    0 : "-ON$\n",
    1 : "qON$\n",
    2 : "wON$\n",
    3 : "eON$\n",
    4 : "rON$\n",
    5 : "tON$\n",
    6 : "yON$\n",
    7 : "uON$\n", 
}

def bytes_to_int(data):
    sign = str(data[:-1])[2:-1]
    if sign in convert_char_to_int:
        convert = convert_char_to_int[sign]
    else:
        convert = str(data[:-1])[4:-1]
        convert = int(convert, 16)
    return convert

def analog_to_lux(intensity):
    REF_RESISTANCE = 5000 # 5k ohm
    LUX_CALC_SCALAR = 12518931
    LUX_CALC_EXPONENT = -1.405
    
    raw = intensity * 4 # return 8 bit value to 10 bit value
    resistorVoltage = raw / 1023 * 5
    ldrVoltage = 5 - resistorVoltage
    ldrResistance = ldrVoltage/resistorVoltage * REF_RESISTANCE
    ldrLux = LUX_CALC_SCALAR * math.pow(ldrResistance, LUX_CALC_EXPONENT)
    return int(ldrLux)

# -------------------------- Machine learning Model Classification ----------------------
def classified(X):
    result = []
    load_rf = joblib.load("./ML_model/random_forest.joblib")
    load_knn = joblib.load("./ML_model/knn.joblib")
    result.append(load_rf.predict(X)[0])
    result.append(load_knn.predict(X)[0])
    return result

# ----------------------------- Excel Database Update ------------------------------------
def update_db(data, res):
    year_now = str(datetime.now().year)
    month_now = str(datetime.now().month)
    day_now = str(datetime.now().day)
    path = "logger/"
    date_now = day_now + "-" + month_now + "-" + year_now
    file_name = "logger/" + date_now + '_logger.xlsx'
    if os.path.exists(file_name):
        wb = load_workbook(filename=file_name)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
    sheet['A1'] = "Datetime"
    sheet['B1'] = "Intensitas Cahaya"
    sheet['C1'] = "Kelembaban"
    sheet['D1'] = "temperature"
    sheet['E1'] = "Hasil Klasifikasi Random Forest"
    sheet['F1'] = "Hasil Klasifikasi KNN"
    max = sheet.max_row
    
    # Data prepared before update
    date_update = str(datetime.now())[:19]
    rows = [data[0], data[1], data[2], res[0], res[1]]
    
    # update data
    sheet.cell(row= 1 + max, column=1).value = date_update
    for c, value in enumerate(rows, start=1):
        sheet.cell(row= 1 + max, column=c+1).value = value
    
    #save the file
    wb.save(filename=file_name)
    wb.close()
    
# ----------------------------- IoT and microcontroller communication ------------------------------------
isDifferent = False
count_classification = 0 # classification flag
received_flag = False
data = [0,0,0]
result = [0,0]

def request():
    global stopEvent
    stopEvent = threading.Event()
    thread = threading.Thread(target = get_data)
    thread.start() # start multithread process
    
def get_data():
    global recheck, count_classification, data, result, isDifferent, recheck
    start = time.perf_counter() # start timer
    var = "?luxData$\n"
    ser.write(var.encode())
    rx_data = ser.read_until('\n')
    lux_value = bytes_to_int(rx_data)
    lux_value = analog_to_lux(lux_value)
    time.sleep(0.5)
    
    var = "?humData$\n"
    ser.write(var.encode())
    rx_data = ser.read_until('\n')
    humidity_value = bytes_to_int(rx_data)
    time.sleep(0.5)
    
    var = "?temData$\n"
    ser.write(var.encode())
    rx_data = ser.read_until('\n')
    tempC_value = bytes_to_int(rx_data)
    time.sleep(0.5)
    
    print("\nlux = " + str(lux_value) + "\nhumidity = " + str(humidity_value) + "\nwater_temp = " + str(tempC_value))
    stop = time.perf_counter() # stop timer
    print("\n\n-------------- Total get sensor data via serial time : " + str(round(stop-start, 2)) + " seconds-------\n")
    
    x = np.array([[lux_value, humidity_value, tempC_value]])
    data = [lux_value, humidity_value, tempC_value]
    input_x = pd.DataFrame(x, columns = ['Suhu(*C)', 'Kelembaban_Udara(%)', 'Intensitas_Cahaya(Lux)'])
    
    # classification section
    result = classified(input_x)
    
    # Checking result
    print(convert_res[result[model]][0])
    print(convert_res[result[model]][1])
    
    start = time.perf_counter() # start timer
    #publishing knn result to web
    publish(x, result[model])
    stop = time.perf_counter() # stop timer
    print("\n\n-------------- Total publish sensor data and classification via MQTT time : " + str(round(stop-start, 2)) + " seconds-------\n")
    
    #update data to db
    update_db(x[0], result) 
    
    #Send knn classification result for output driving purposed
    result_send = res_to_out[result[model]]
    ser.write(result_send.encode())
    ser.flushInput()
    
    stopEvent.set() # end multithread process

def publish(data, res):
    condition_value = convert_res[res][0]
    note_value = convert_res[res][1]
    aio.send_data(lux, int(data[0][0]))
    aio.send_data(humidity, int(data[0][1]))
    aio.send_data(temp, int(data[0][2]))
    aio.send_data(update_condition, condition_value)    
    aio.send_data(note, note_value)  
    print("Data berhasil terkirim")

previous_onDuration = 0
previous_offDuration = 0
previous_lampState = '0'
previous_pumpState = '0'
different_duration = False
different_lampState = False
different_pumpState = False

def subscribe():
    global previous_onDuration, previous_offDuration, previous_lampState, previous_pumpState, different_lampState, different_pumpState, different_duration
    start = time.perf_counter() # start timer
    lampfeed = aio.receive(lamp_state).value
    pumpfeed = aio.receive(pump_state).value
    onDuration = aio.receive(pump_onDuration).value
    offDuration = aio.receive(pump_offDuration).value
    print("lamp feed = {0}\npump feed = {1}\nonDuration_feed = {2}\noffDuration_feed = {3}".format(lampfeed, pumpfeed, onDuration, offDuration))
    stop = time.perf_counter() # stop timer
    print("\n\n-------------- Total getting IO control value from web via MQTT time : " + str(round(stop-start, 2) + subscribe_interval) + " seconds-------\n")
    if lampfeed == '1':
        lamp_state_send = "dON$\n"
    if lampfeed == '0':
        lamp_state_send = "dOFF$\n"
    if pumpfeed == '1':
        pump_state_send = "oON$\n"
    if pumpfeed == '0':
        pump_state_send = "oOFF$\n"
    
    # preventing repetition
    if onDuration != previous_onDuration:
        previous_onDuration = onDuration
        different_duration = True
    if offDuration != previous_offDuration:
        previous_offDuration = offDuration
        different_duration = True
    if pumpfeed != previous_pumpState:
        previous_pumpState = pumpfeed
        different_pumpState = True
    if lampfeed != previous_lampState:
        previous_lampState = lampfeed
        different_lampState = True
    
    # gate 
    if different_duration == True:
        duration_update = "p," + str(onDuration) + "," + str(offDuration) + "$\n"
        ser.write(duration_update.encode())
        ser.flushInput()
        different_duration = False
    if different_lampState == True:
        ser.write(lamp_state_send.encode())
        ser.flushInput()
        different_lampState = False
    if different_pumpState == True:
        ser.write(pump_state_send.encode())
        ser.flushInput()
        different_pumpState = False

# ------------------------------------------------------------------------------------------
# --------------------------------------------- GUI Section --------------------------------

white       = "#ffffff"
BlackSolid  = "#000000"
font        = "Constantia"
fontButtons = (font, 12)
maxWidth    = 640
maxHeight   = 480 

def _from_rgb(rgb):
    """translate an rgb tuple to hex"""
    return "#%02x%02x%02x" % rgb

# vanilla button class
class buttonL:
    def __init__(self, obj, size, position, text,font, fontSize, hoverColor,command=None):
        self.obj= obj
        self.size= size
        self.position= position
        self.font= font
        self.fontSize= fontSize
        self.hoverColor= hoverColor
        self.text= text
        self.command = command
        self.state = True
        self.Button_ = None

    def myfunc(self):
        print("Hello size :" , self.size)
        print("Hello position :" , self.position)
        print("Hello font :" , self.font)
        print("Hello fontSize :" , self.fontSize)
        print("Hello hoverState :" , self.hoverColor)
  
    def changeOnHover(self, obj,colorOnHover, colorOnLeave):
         obj.bind("<Enter>", func=lambda e: obj.config(
             background=colorOnHover))

         obj.bind("<Leave>", func=lambda e: obj.config(
             background=colorOnLeave))
            
    def buttonShow(self):
        fontStyle = tkFont.Font(family= self.font, size=self.fontSize,weight="bold")
        self.Button_ = Button(self.obj,text = self.text, font=fontStyle, width = self.size[0], height = self.size[1],  bg = self.hoverColor[1] if isinstance(self.hoverColor, list)  == True else  self.hoverColor, compound=TOP,command=self.command)         
        self.Button_.place(x=self.position[0],y=self.position[1])

        if isinstance(self.hoverColor, list) == True:
            self.changeOnHover(self.Button_, self.hoverColor[0], self.hoverColor[1])
        else:
            self.changeOnHover(self.Button_, self.hoverColor, self.hoverColor)
    
    def stateButton(self,st):
        self.st=st
        if not self.Button_ == None:
            self.Button_["state"]=self.st
    
    def buttonUpdate(self, textUpdate = "", colorUpdate = "#fff"):
        temp= [self.hoverColor[0], colorUpdate]
        self.hoverColor = temp
        self.Button_.config(text = textUpdate, bg = self.hoverColor[1] if isinstance(self.hoverColor, list)  == True else  self.hoverColor)
        if isinstance(self.hoverColor, list) == True:
            self.changeOnHover(self.Button_, self.hoverColor[0], self.hoverColor[1])
        else:
            self.changeOnHover(self.Button_, self.hoverColor, self.hoverColor)

# image button class
class buttonImg:
    def __init__(self, obj, imgDir, size, position, hoverColor, command=None):
        self.obj= obj
        self.imgDir= imgDir
        self.size= size
        self.position= position
        self.hoverColor = hoverColor
        self.command = command
        self.state = True
        self.Button_ = None
    
    def changeOnHover(self, obj,colorOnHover, colorOnLeave):
         obj.bind("<Enter>", func=lambda e: obj.config(
             background=colorOnHover))

         obj.bind("<Leave>", func=lambda e: obj.config(
             background=colorOnLeave))
         
    def buttonShow(self):
        self.Button_ = Button(self.obj, width = self.size[0], height = self.size[1], bg = self.hoverColor[1] if isinstance(self.hoverColor, list) == True else self.hoverColor, bd = 10, highlightthickness=4, highlightcolor="#000", highlightbackground="#000", borderwidth = 4, compound=TOP, command=self.command)         
        self.Button_.place(x=self.position[0],y=self.position[1])
        self.imageOpen = Image.open(self.imgDir)
        self.imageOpen = self.imageOpen.resize((self.size[0],self.size[1]), Image.ANTIALIAS)
        self.imageOpen = ImageTk.PhotoImage(self.imageOpen)
        self.Button_.config(image=self.imageOpen)
        
        if isinstance(self.hoverColor, list) == True:
            self.changeOnHover(self.Button_, self.hoverColor[0], self.hoverColor[1])
        else:
            self.changeOnHover(self.Button_, self.hoverColor, self.hoverColor)
    
    def stateButton(self,st):
        self.st=st
        if not self.Button_ == None:
            self.Button_["state"]=self.st
    
    def buttonUpdate(self, colorUpdate = "#fff"):
        temp= [self.hoverColor[0], colorUpdate]
        self.hoverColor = temp
        self.Button_.config(bg = self.hoverColor[1] if isinstance(self.hoverColor, list)  == True else  self.hoverColor)
        if isinstance(self.hoverColor, list) == True:
            self.changeOnHover(self.Button_, self.hoverColor[0], self.hoverColor[1])
        else:
            self.changeOnHover(self.Button_, self.hoverColor, self.hoverColor)

class logo:
    def __init__(self, obj, imgDir, size, position, bg, command=None):
        self.obj = obj
        self.imgDir = imgDir
        self.size = size
        self.position = position
        self.bg = bg
        self.command = command
        self.state = True
        self.Button_ = None
        
    def show(self):
        self.logo = Button(self.obj, width = self.size[0], height = self.size[1], bg = self.bg, borderwidth = 0)
        self.logo.place(x = self.position[0], y = self.position[1])
        self.img = Image.open(self.imgDir)
        self.img = self.img.resize((self.size[0], self.size[1]),  Image.ANTIALIAS)
        self.img = ImageTk.PhotoImage(self.img)
        self.logo.config(image = self.img)
                           
class framecontroller(tk.Tk):
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        
        #Graphics window
        self.mainWindow = self
        self.mainWindow.configure(bg=BlackSolid)
        self.mainWindow.geometry('%dx%d+%d+%d' % (maxWidth,maxHeight,0,0))
        self.mainWindow.resizable(0,0)
        self.mainWindow.title("Smart Agriculture")
        self.mainWindow.attributes("-fullscreen", True)
        
        # # creating a container
        container = tk.Frame(self.mainWindow) 
        container.configure(bg=BlackSolid)
        container.pack(side = "top", fill = "both", expand = True)
  
        container.grid_rowconfigure(0, weight = 1)
        container.grid_columnconfigure(0, weight = 1)
        
        frame = StartPage(container, self.mainWindow)
        frame.grid(row = 0, column = 0, sticky = 'nsew')
        frame.tkraise()
        
class StartPage(tk.Frame):
    def __init__(self, parent, controller):
        self.waterPump_flag = False
        self.peltier_flag = False
        self.lamp_flag = False
        self.humidifier_flag = False
        self.bg = "#ddd"
        self.fg = "#000"
        
        tk.Frame.__init__(self, parent)
        self.parent = parent
        # backgroud
        self.configure(bg=self.bg)
        
        # Logo
        self.ub = logo(self, 'Picture2.png', [50, 50], [10, 10], bg = self.bg)
        self.ub.show()
        
        self.electro = logo(self, 'Picture1.png', [50, 50], [70, 10], bg = self.bg)
        self.electro.show()
        
        # contain
        # Showing each sensor value
        fontStyleLabel= tkFont.Font(family="Arial", size=55, weight = "bold")
        
        #self.condLabel = Label(self, text="Tidak Ideal", bg='#444', fg='#fff', font=fontStyleLabel)
        self.condLabel = Label(self, text="      Ideal", bg=self.bg, fg=self.fg, font=fontStyleLabel)
        self.condLabel.pack()
        self.condLabel.place(x=120,y=46)
        
        fontStyleLabel= tkFont.Font(family="Arial", size=25)
        
        self.tempLabel = Label(self, text="Temperature", bg=self.bg , fg=self.fg, font=fontStyleLabel)
        self.tempLabel.pack()
        self.tempLabel.place(x=20,y=150)
        
        fontStyleLabel= tkFont.Font(family="Arial", size=45)
        self.tempValue = Label(self, text="0", bg=self.bg, fg=self.fg, font=fontStyleLabel)
        self.tempValue.pack()
        self.tempValue.place(x=20,y=190)
        
        fontStyleLabel= tkFont.Font(family="Arial", size=25)
        self.humLabel = Label(self, text="Kelembapan", bg=self.bg, fg=self.fg, font=fontStyleLabel)
        self.humLabel.pack()
        self.humLabel.place(x=400,y=150)
        
        fontStyleLabel= tkFont.Font(family="Arial", size=45)
        self.humValue = Label(self, text="0", bg=self.bg, fg=self.fg, font=fontStyleLabel)
        self.humValue.pack()
        self.humValue.place(x=400,y=190)
        
        fontStyleLabel= tkFont.Font(family="Arial", size=25)
        self.lightLabel = Label(self, text="Intensitas Cahaya", bg=self.bg, fg=self.fg, font=fontStyleLabel)
        self.lightLabel.pack()
        self.lightLabel.place(x=180,y=275)
        
        fontStyleLabel= tkFont.Font(family="Arial", size=45)
        self.lightValue = Label(self, text="0", bg=self.bg, fg=self.fg, font=fontStyleLabel)
        self.lightValue.pack()
        self.lightValue.place(x=180,y=315)
        
        # Showing each unit sensor values
        fontStyleLabel= tkFont.Font(family="Arial", size=45)
        self.tempUnit = Label(self, text="*C", bg=self.bg, fg=self.fg, font=fontStyleLabel)
        self.tempUnit.pack()
        self.tempUnit.place(x=180,y=190)
        
        self.humUnit = Label(self, text="%", bg=self.bg, fg=self.fg, font=fontStyleLabel)
        self.humUnit.pack()
        self.humUnit.place(x=570,y=190)
        
        self.lightUnit = Label(self, text="Lux", bg=self.bg, fg=self.fg, font=fontStyleLabel)
        self.lightUnit.pack()
        self.lightUnit.place(x=370,y=315)
        
        # Actuator manually control button 
        fontStyle = tkFont.Font(family= "Arial", size=25,weight="bold")
        
        self.button = buttonL(self,[8,2],[5,400],"Pompa air",fontStyle,18,[BlackSolid,_from_rgb((244,239,140))],lambda : [self.waterPump()])
        self.button.buttonShow()
        
        self.button2 = buttonL(self,[7,2],[185,400],"Lampu",fontStyle,18,[BlackSolid,_from_rgb((255,190,100))],lambda : [self.lamp()])
        self.button2.buttonShow()
        
        self.button3 = buttonL(self,[7,2],[338,400],"Peltier",fontStyle,18,[BlackSolid,_from_rgb((255,190,100))],lambda : [self.peltier()])
        self.button3.buttonShow()
        
        self.button4 = buttonL(self,[7,2],[490,400],"Humidifier",fontStyle,18,[BlackSolid,_from_rgb((255,190,100))],lambda : [self.humidifier()])
        self.button4.buttonShow()
        
        # GUI auto update
        self.parent.after(2000, self.update_gui)

    def update_gui(self):
        self.lightValue.configure(text=str(data[0]))
        self.humValue.configure(text=str(data[1]))
        self.tempValue.configure(text=str(data[2]))
        if(result[0] == 1):
            self.condLabel.configure(text = "      Ideal")
        else:
            self.condLabel.configure(text = "Tidak Ideal")
        self.parent.after(2000, self.update_gui)
        
    def waterPump(self, event = None):
        self.waterPump_flag = not(self.waterPump_flag)
        self.start = time.perf_counter()
        print(self.waterPump_flag)
        if(self.waterPump_flag):
            pump_state_send = "oON$\n"
            ser.write(pump_state_send.encode())
            self.stop = time.perf_counter()
            print("\n\n-------------- Total IO control via serial time : " + str(round(self.stop-self.start, 2)) + " seconds-------\n")
            messagebox.showinfo("notification", "Pompa Air Menyala !!")
        else:
            pump_state_send = "oOFF$\n"
            ser.write(pump_state_send.encode())
            self.stop = time.perf_counter()
            print("\n\n-------------- Total IO control via serial time : " + str(round(self.stop-self.start, 2)) + " seconds-------\n")
            messagebox.showinfo("notification", "Pompa Air Mati !!")
    
    def lamp(self, event = None):
        self.lamp_flag = not(self.lamp_flag)
        self.start = time.perf_counter()
        print(self.lamp_flag)
        if(self.lamp_flag):
            lamp_state_send = "dON$\n"
            ser.write(lamp_state_send.encode())
            self.stop = time.perf_counter()
            print("\n\n-------------- Total IO control via serial time : " + str(round(self.stop-self.start, 2)) + " seconds-------\n")
            messagebox.showinfo("notification", "Lamu UV Menyala !!")
        else:
            lamp_state_send = "dOFF$\n"
            ser.write(lamp_state_send.encode())
            self.stop = time.perf_counter()
            print("\n\n-------------- Total IO control via serial time : " + str(round(self.stop-self.start, 2)) + " seconds-------\n")
            messagebox.showinfo("notification", "Lampu UV Mati !!")
        
    def peltier(self, event = None):
        self.peltier_flag = not(self.peltier_flag)
        self.start = time.perf_counter()
        print(self.peltier_flag)
        if(self.peltier_flag):
            peltier_state_send = "aON$\n"
            ser.write(peltier_state_send.encode())
            self.stop = time.perf_counter()
            print("\n\n-------------- Total IO control via serial time : " + str(round(self.stop-self.start, 2)) + " seconds-------\n")
            ser.flushInput()
            messagebox.showinfo("notification", "Peltier Menyala !!")
        else:
            peltier_state_send = "aOFF$\n"
            ser.write(peltier_state_send.encode())
            self.stop = time.perf_counter()
            print("\n\n-------------- Total IO control via serial time : " + str(round(self.stop-self.start, 2)) + " seconds-------\n")
            ser.flushInput()
            messagebox.showinfo("notification", "Peltier Mati !!")
        
    def humidifier(self, event = None):
        self.humidifier_flag = not(self.humidifier_flag)
        self.start = time.perf_counter()
        print(self.humidifier_flag)
        if(self.humidifier_flag):
            humidifier_state_send = "sON$\n"
            ser.write(humidifier_state_send.encode())
            self.stop = time.perf_counter()
            print("\n\n-------------- Total IO control via serial time : " + str(round(self.stop-self.start, 2)) + " seconds-------\n")
            messagebox.showinfo("notification", "Humidifier Menyala !!")
        else:
            humidifier_state_send = "sOFF$\n"
            ser.write(humidifier_state_send.encode())
            self.stop = time.perf_counter()
            print("\n\n-------------- Total IO control via serial time : " + str(round(self.stop-self.start, 2)) + " seconds-------\n")
            messagebox.showinfo("notification", "Humidifier Mati !!")
        
# -------------------------------------- Program Execution ------------------------------
if __name__ == '__main__':
    app = framecontroller()
    scheduler = BackgroundScheduler()
    scheduler.add_job(request, 'interval', seconds = request_interval)
    scheduler.add_job(subscribe, "interval", seconds = subscribe_interval)
    scheduler.start()
    print('Press Ctrl+{0} to exit'.format('Break' if os.name == 'nt' else 'C'))    
    try:
        # This is here to simulate application activity (which keeps the main thread alive)
        while True:
            app.mainloop()
    except:
        # Not strictly necessary if daemonic mode is enabled but should be done if possible
        scheduler.shutdown()

    

        