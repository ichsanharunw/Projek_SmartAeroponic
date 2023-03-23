from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import random
import os

# Create file
year_now = str(datetime.now().year)
month_now = str(datetime.now().month)
day_now = str(datetime.now().day)
path = "logger/"
date_now = day_now + "-" + month_now + "-" + year_now
file_name = "logger/" + date_now + '_logger.xlsx'
if os.path.exists(file_name):
    wb = load_workbook(filename=file_name)
    sheet = wb.active
else:
    wb = Workbook()
    sheet = wb.active

#sheet = workbook.active
count = 0
rows = []
sheet['A1'] = "Datetime"
sheet['B1'] = "Kelembaban"
sheet['C1'] = "Intensitas Cahaya"
sheet['D1'] = "temperature"
max = sheet.max_row

while True:
    date_update = str(datetime.now())[:19]
    humidity_value = random.randint(0, 100)
    lux_value = random.randint(0, 255)
    temp_value = random.randint(0, 100)
    rows = [humidity_value, lux_value, temp_value]
    sheet.cell(row= 1 + max, column=1).value = date_update
    for c, value in enumerate(rows, start=1):
        sheet.cell(row= 1 + max, column=c+1).value = value
    count += 1
    if(count > 15):
        count = 0
        break


print("Selesai Lur")

#save the file
wb.save(filename=file_name)
wb.close()